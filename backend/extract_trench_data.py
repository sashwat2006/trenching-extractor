# Only keep Excel writing and file handling logic here. All parser-specific logic is now in their respective files
from parsers.mcgm import (
    non_refundable_request_parser as mcgm_non_refundable_parser, 
    sd_parser as mcgm_sd_parser, 
    HEADERS, STATIC_VALUES
)
from parsers.mbmc import (
    non_refundable_request_parser as mbmc_non_refundable_parser,
    sd_parser as mbmc_sd_parser
)
import openpyxl
import re
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import camelot

# --- Excel Writing Logic ---
def append_row_to_excel(excel_path, row, headers, manual_fields=None, blue_headers=None):
    import os
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    print(f"[DEBUG] Writing Excel file to: {os.path.abspath(excel_path)}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    ws.append(row)
    # Style (generic, no MCGM-specific logic)
    header_font = Font(name='Calibri', size=11, bold=True)  # Calibri 11, bold
    data_font = Font(name='Calibri', size=11)  # Calibri 11
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    blue_fill = PatternFill(start_color="B7E1FA", end_color="B7E1FA", fill_type="solid")  # Light blue
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    blue_headers_set = set(blue_headers or [])
    # Header row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        if headers[col-1] in blue_headers_set:
            cell.fill = blue_fill
        else:
            cell.fill = header_fill
        cell.alignment = center_wrap
        cell.border = border
    # Data row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = data_font
        cell.alignment = center_wrap
        cell.border = border
    # Adjust column widths and wrap text
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    wb.save(excel_path)

def process_demand_note(uploaded_file_path, authority, manual_values=None, sd_manual_values=None, return_paths=False):
    """
    Handles the uploaded PDF and returns paths to the generated Excel files for non-refundable and SD outputs.
    Also returns a flag if the majority of dynamic fields are blank (for file validation in UI).
    Accepts manual_values dict for MCGM non-refundable blue-highlighted fields and sd_manual_values for SD output blue fields.
    """
    import os
    import fitz
    print(f"[DEBUG] process_demand_note: uploaded_file_path={uploaded_file_path}, authority={authority}")
    tmp_pdf_path = uploaded_file_path  # Now just use the path directly
    base, _ = os.path.splitext(tmp_pdf_path)

    # Define blue/manual headers for MCGM
    blue_headers_non_ref = [
        "LM/BB/FTTH", "GO RATE", "Total Route (MTR)", "Not part of capping (License Fee/Rental Payment /Way Leave charges etc.)",
        "REASON FOR DELAY (>2 DAYS)", "PO No.", "Route Name(As per CWIP)", "Section Name for ROW(As per CWIP)"
    ]
    blue_headers_sd = [
        "Execution Partner GBPA PO No.", "Partner PO circle", "Unique route id", "NFA no."
    ]
    # Define blue/manual headers for MBMC (fill this list as needed)
    blue_headers_non_ref_mbmc = [
        "LM/BB/FTTH", "GO RATE", "Total Route (MTR)", "Not part of capping (License Fee/Rental Payment /Way Leave charges etc.)",
        "REASON FOR DELAY (>2 DAYS)", "PO No.", "Route Name(As per CWIP)", "Section Name for ROW(As per CWIP)"
        # Add more MBMC manual headers here as needed
    ]
    blue_headers_sd_mbmc = [
        "Execution Partner GBPA PO No.", "Partner PO circle", "Unique route id", "NFA no."
        # Add more MBMC SD manual headers here as needed
    ]
    # Non-refundable output
    if authority.upper() == "MCGM":
        row = mcgm_non_refundable_parser(tmp_pdf_path, manual_values=manual_values)
        print(f"[DEBUG] [excel] Writing row to Non-Refundable Excel: {row}")
        # Re-extract demand note number after manual fields are applied
        try:
            demand_note_number = row[HEADERS.index("Demand Note Reference number")]
        except Exception:
            demand_note_number = "UnknownDemandNote"
        if not demand_note_number:
            demand_note_number = "UnknownDemandNote"
        safe_demand_note_number = sanitize_filename(demand_note_number)
        tmp_xlsx_path = os.path.join(os.path.dirname(base), f"{safe_demand_note_number}_Non Refundable Output.xlsx")
        append_row_to_excel(tmp_xlsx_path, row, HEADERS, manual_fields=manual_values, blue_headers=blue_headers_non_ref)
        # SD output for MCGM
        alt_headers, row_alt = mcgm_sd_parser(tmp_pdf_path, manual_values=sd_manual_values)
        tmp_xlsx_alt_path = os.path.join(os.path.dirname(base), f"{safe_demand_note_number}_SD Output.xlsx")
        append_row_to_excel(tmp_xlsx_alt_path, row_alt, alt_headers, manual_fields=sd_manual_values, blue_headers=blue_headers_sd)
        sd_xlsx_alt_path = tmp_xlsx_alt_path
    elif authority.upper() == "MBMC":
        row = mbmc_non_refundable_parser(tmp_pdf_path, manual_values=manual_values)
        print(f"[DEBUG] [excel] Writing row to Non-Refundable Excel: {row}")
        try:
            demand_note_number = row[HEADERS.index("Demand Note Reference number")]
        except Exception:
            demand_note_number = "UnknownDemandNote"
        if not demand_note_number:
            demand_note_number = "UnknownDemandNote"
        safe_demand_note_number = sanitize_filename(demand_note_number)
        tmp_xlsx_path = os.path.join(os.path.dirname(base), f"{safe_demand_note_number}_Non Refundable Output.xlsx")
        append_row_to_excel(tmp_xlsx_path, row, HEADERS, manual_fields=manual_values, blue_headers=blue_headers_non_ref_mbmc)
        # SD output for MBMC
        alt_headers, row_alt = mbmc_sd_parser(tmp_pdf_path, manual_values=sd_manual_values)
        tmp_xlsx_alt_path = os.path.join(os.path.dirname(base), f"{safe_demand_note_number}_SD Output.xlsx")
        append_row_to_excel(tmp_xlsx_alt_path, row_alt, alt_headers, manual_fields=sd_manual_values, blue_headers=blue_headers_sd_mbmc)
        sd_xlsx_alt_path = tmp_xlsx_alt_path
    else:
        row = mcgm_non_refundable_parser(tmp_pdf_path)
        print(f"[DEBUG] [excel] Writing row to Non-Refundable Excel: {row}")
        try:
            demand_note_number = row[HEADERS.index("Demand Note Reference number")]
        except Exception:
            demand_note_number = "UnknownDemandNote"
        if not demand_note_number:
            demand_note_number = "UnknownDemandNote"
        safe_demand_note_number = sanitize_filename(demand_note_number)
        tmp_xlsx_path = os.path.join(os.path.dirname(base), f"{safe_demand_note_number}_Non Refundable Output.xlsx")
        append_row_to_excel(tmp_xlsx_path, row, HEADERS, manual_fields=manual_values, blue_headers=blue_headers_non_ref)
        sd_xlsx_alt_path = None
    # Check if majority of dynamic fields are blank (only those present in HEADERS)
    dynamic_fields = [
        "Demand Note Reference number",
        "Section Length (Mtr.)",
        "GST Amount",
        "SD Amount",
        "ROW APPLICATION  DATE",
        "Demand Note Date",
        "DN RECEIVED FROM PARTNER/AUTHORITY- DATE",
        "Difference from, DN date  - DN Sent to Central team (ARTL)",
        "Total DN Amount ( NON REFUNDABLE+SD+BG+GST) To be filled by helpdesk team",
        "Road Types - CC/BT/TILES/ Normal Soil/kacha",
        "Rate/mtr- Current DN (UG/OH)",
        "Covered under capping (Restoration Charges, admin, registration etc.)",
        "Not part of capping (License Fee/Rental Payment /Way Leave charges etc.)",
        "Non Refundable Cost( Amount to process for payment shold be sum of 'Z' and 'AA' coulm )"
    ]
    present_fields = [f for f in dynamic_fields if f in HEADERS]
    blank_count = sum(1 for f in present_fields if row[HEADERS.index(f)] == "" or row[HEADERS.index(f)] is None)
    majority_blank = blank_count >= (len(present_fields) // 2 + 1)
    if return_paths:
        print(f"[DEBUG] Returning paths: Non-Refundable: {tmp_xlsx_path}, SD: {sd_xlsx_alt_path}")
        return tmp_xlsx_path, sd_xlsx_alt_path, demand_note_number
    # Default: return the Non-Refundable Excel file as bytes (for FastAPI StreamingResponse)
    with open(tmp_xlsx_path, "rb") as f:
        excel_bytes = f.read()
    # Return both the bytes and the filename for FastAPI to use in Content-Disposition
    return excel_bytes, f"{demand_note_number}_Non Refundable Output.xlsx"

def sanitize_filename(name):
    # Replace all non-alphanumeric and non-underscore/dash with underscore
    return re.sub(r'[^\w\-]', '_', name)