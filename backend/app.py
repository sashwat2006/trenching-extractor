import streamlit as st
import os
import traceback
import openpyxl
from pathlib import Path

from extract_trench_data import process_demand_note

# --- Page Setup ---
st.set_page_config(
    page_title="Municipal Demand Note Parser",
    page_icon="cloudextel_logo.png",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- Styling (Consolidated) ---
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
    html, body, .main, .block-container {
        background: linear-gradient(120deg, #0a1626 0%, #1a2233 100%) !important;
        color: #fff !important;
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif !important;
    }
    .stApp { background: transparent !important; }    .stButton>button, .stDownloadButton>button {
        background: linear-gradient(90deg, #003366 0%, #0052cc 100%);
        color: #fff;
        font-size: 0.95em;
        font-weight: bold;
        border-radius: 8px;
        padding: 0.5em 1.2em;
        margin: 0.3em 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.10);
        border: none;
    }
    .stButton>button:hover, .stDownloadButton>button:hover {
        background: linear-gradient(90deg, #0052cc 0%, #003366 100%);
        transform: scale(1.02);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }
    .download-section {
        display: flex;
        align-items: center;
        gap: 1em;
        margin-bottom: 1em;
    }
    .download-section h3 {
        margin: 0;
        flex-grow: 1;
    }
    .stFileUploader>div>div {
        background: #101c33;
        border: 2px dashed #0052cc;
        border-radius: 14px;
        color: #fff;
        margin-bottom: 1.5em;
    }
    .stDataFrame {
        background: #101c33;
        border-radius: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        color: #fff;
        font-size: 1.05em;
    }
    .stDataFrame thead tr th {
        height: 28px !important;
        min-height: 28px !important;
        max-height: 28px !important;
        font-family: 'Times New Roman', Times, serif !important;
        font-size: 1em !important;
        font-weight: bold !important;
        background: #FFFF00 !important;
        color: #222 !important;
        vertical-align: middle !important;
        text-align: center !important;
        white-space: normal !important;
    }
    .block-container {
        padding-top: 3.5rem;
        padding-bottom: 3.5rem;
    }
    .cloudextel-title {
        font-size: 2.1em;
        font-weight: 700;
        color: #fff;
        letter-spacing: 0.01em;
        line-height: 1.1;
    }
    .cloudextel-caption {
        color: #b0c4de;
        font-size: 1.08em;
        margin-top: 0.7em;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 0 !important;
        padding: 0 2.5em !important;
        margin-left: 2.5em !important;
        margin-bottom: 1.2em !important;
        background: #111 !important; /* jet black */
        border-radius: 0 !important;
        border-bottom: 2px solid #222 !important;
        box-shadow: none !important;
        width: calc(100% - 2.5em) !important;
        max-width: 1600px;
        position: relative;
    }
    .stTabs [data-baseweb="tab"] {
        font-size: 1em;
        font-weight: 500;
        border-radius: 0 !important;
        padding: 0.7em 2.5em !important;
        color: #bbb;
        background: #111 !important;
        border: none !important;
        box-shadow: none !important;
        margin-bottom: 0 !important;
        transition: background 0.2s, color 0.2s;
        flex: 1 1 0;
        text-align: center;
        min-width: 0;
    }
    .stTabs [aria-selected="true"] {
        background: #000 !important; /* jet black for selected tab */
        color: #fff !important;
        border-bottom: 2px solid #fff !important;
        z-index: 2;
    }
    .stTabs [data-baseweb="tab"]:hover {
        background: #222 !important;
        color: #fff !important;
    }
    .stDownloadButton>button:hover {
        color: #fff !important;
    }
    html, body, [data-testid="stAppViewContainer"] {
        height: 100%;
        min-height: 100vh;
        background: #10192b !important;
    }
    [data-testid="stAppViewContainer"] > .main {
        min-height: 100vh;
        background: #10192b !important;
        padding-bottom: 0 !important;
    }
    [data-testid="stSidebar"] {
        background: #10192b !important;
    }
    </style>
""", unsafe_allow_html=True)

# --- Header ---
col1, col2 = st.columns([0.18, 0.82])
with col1:
    st.image("cloudextel_logo.png", width=110, use_container_width=True)
with col2:
    st.markdown("""
        <div style='margin-left: 2.5em;'>
            <div class='cloudextel-title'>Municipal Demand Note Parser</div>
            <div class='cloudextel-caption'>NLP-powered extraction for Regional trenching Demand Notes</div>
        </div>
    """, unsafe_allow_html=True)

st.markdown("""
    <hr style='border: 1px solid #223355; margin: 0.5em 0 2em 0;'/>
""", unsafe_allow_html=True)

# --- Tabbed Uploads ---
st.markdown("""
    <div style='margin-bottom: 2em;'></div>
""", unsafe_allow_html=True)
authorities = ["KDMC", "MBMC", "MCGM", "MIDC Type 1", "MIDC Type 2", "NMMC"]
tabs = st.tabs(authorities)

def display_output_section(title, data_path, authority_type, is_sd=False, demand_number=None):
    """
    Helper function to display a section with data and download button.
    Args:
        title: String title for the section
        data_path: Path to the Excel file
        authority_type: Type of authority (for filename)
        is_sd: Whether this is an SD output section
        demand_number: The extracted demand note number (for filename)
    """
    import openpyxl
    import os
    import traceback
    try:
        st.markdown(f"""
            <div class="download-section">
                <h3>{title}</h3>
            </div>
        """, unsafe_allow_html=True)
        col1, col2 = st.columns([0.85, 0.15])
        with col1:
            try:
                wb = openpyxl.load_workbook(data_path)
                ws = wb.active
                data = list(ws.values)
                headers = data[0]
                row = data[1] if len(data) > 1 else []
                st.dataframe({h: [v] for h, v in zip(headers, row)}, use_container_width=True)
                wb.close()
            except Exception as e:
                st.error(f"Error reading Excel file: {str(e)}")
                return
        with col2:
            try:
                suffix = "_sd output" if is_sd else "_Non Refundable Request"
                # Use demand_number if available, else fallback to authority_type
                if demand_number:
                    filename = f"{demand_number}{suffix}.xlsx"
                else:
                    filename = f"cloudextel_output_{authority_type.replace(' ', '_')}{suffix}.xlsx"
                with open(data_path, 'rb') as f:
                    st.download_button(
                        label="Download Client File",
                        data=f.read(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Error preparing download: {str(e)}")
        st.markdown("---")
    except Exception as e:
        st.error(f"An error occurred: {e}")
        st.code(traceback.format_exc(), language='python')
    finally:
        for f in [locals().get("tmp_pdf_path"), locals().get("tmp_xlsx_path"), locals().get("tmp_xlsx_alt_path")]:
            if f and os.path.exists(f):
                try:
                    os.remove(f)
                except:
                    pass

for i, tab in enumerate(tabs):
    with tab:
        st.markdown(f"<div style='font-size:1.3em; font-weight:600; color:#fff; margin-bottom:0.5em;'>Upload {authorities[i]} Demand Note PDF</div>", unsafe_allow_html=True)
        uploaded_file = st.file_uploader(f"Upload a {authorities[i]} PDF", type=["pdf"], key=f"uploader_{i}")
        manual_fields = [
            "LM/BB/FTTH",
            "GO RATE",
            "Total Route (MTR)",
            "Not part of capping (License Fee/Rental Payment /Way Leave charges etc.)",
            "REASON FOR DELAY (>2 DAYS)",
            "PO No.",
            "Route Name(As per CWIP)",
            "Section Name for ROW(As per CWIP)"
        ]
        sd_manual_fields = [
            "Execution Partner GBPA PO No.",
            "Partner PO circle",
            "Unique route id",
            "NFA no."
        ]
        manual_values = {}
        sd_manual_values = {}
        if authorities[i] == "MCGM" and uploaded_file:
            st.markdown("""
                <div style='margin-bottom:0.5em; color:#b0c4de; font-size:1.1em;'><b>Enter values for Non-Refundable Request (Blue fields):</b></div>
            """, unsafe_allow_html=True)
            cols = st.columns(2)
            for idx, field in enumerate(manual_fields):
                with cols[idx % 2]:
                    manual_values[field] = st.text_input(field, key=f"manual_{field}")
            st.markdown("""
                <div style='margin:1.5em 0 0.5em 0; color:#b0c4de; font-size:1.1em;'><b>Enter values for SD Output (Blue fields):</b></div>
            """, unsafe_allow_html=True)
            cols_sd = st.columns(2)
            for idx, field in enumerate(sd_manual_fields):
                with cols_sd[idx % 2]:
                    sd_manual_values[field] = st.text_input(field, key=f"sd_manual_{field}")
        if uploaded_file:
            status_placeholder = st.empty()
            status_placeholder.info("Parsing and extracting data... Please wait.")
            try:
                # Process the file using the backend controller
                result = process_demand_note(uploaded_file, authorities[i], manual_values if authorities[i]=="MCGM" else None, sd_manual_values if authorities[i]=="MCGM" else None)
                if isinstance(result, tuple) and len(result) == 3:
                    non_refundable_path, sd_path, majority_blank = result
                else:
                    non_refundable_path, sd_path = result
                    majority_blank = False
                # Extract demand number for filename (only for MCGM)
                demand_number = None
                if authorities[i] == "MCGM" and non_refundable_path:
                    import openpyxl
                    wb = openpyxl.load_workbook(non_refundable_path)
                    ws = wb.active
                    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                    if "Demand Note Reference number" in headers:
                        idx = headers.index("Demand Note Reference number")
                        demand_number = row[idx] if idx < len(row) else None
                    wb.close()
                # Show warning if majority of dynamic fields are blank (for MCGM)
                if authorities[i] == "MCGM" and majority_blank:
                    status_placeholder.warning("This file does not appear to be a valid MCGM Demand Note. The majority of dynamic fields could not be extracted. Please check your upload.")
                else:
                    status_placeholder.success("File parsed successfully!")
                # Display non-refundable request output
                if non_refundable_path:
                    display_output_section("Non Refundable Request Output", non_refundable_path, authorities[i], is_sd=False, demand_number=demand_number)
                # Display SD output if available
                if sd_path:
                    st.markdown("---")
                    display_output_section("SD Output", sd_path, authorities[i], is_sd=True, demand_number=demand_number)
            except NotImplementedError:
                status_placeholder.info("")
                st.markdown("""
                    <div style='margin-top:2em; font-size:1.15em; color:#b0c4de;'>
                        <b>Parser coming soon!</b> This authority's extraction logic will be available in a future update.
                    </div>
                """, unsafe_allow_html=True)
            except Exception as e:
                status_placeholder.error(f"An error occurred: {str(e)}")
                st.code(traceback.format_exc(), language='python')
            finally:
                # Clean up temporary files
                for path in [locals().get(p) for p in ['non_refundable_path', 'sd_path']]:
                    if path and os.path.exists(path):
                        try:
                            os.remove(path)
                        except:
                            pass
        else:
            st.info(f"Please upload a {authorities[i]} PDF to begin.")